import os
import django
import openpyxl

# Configurar entorno Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "app_stock.settings")
django.setup()

from stock.models import Material

archivo_excel = 'stock_sap.xlsx'
wb = openpyxl.load_workbook(archivo_excel)
hoja = wb.active

# Leer cabeceras de la primera fila
for fila in hoja.iter_rows(min_row=1, max_row=3, values_only=True):
    if fila and "Número de material" in [str(c).strip() for c in fila if c]:
        cabeceras = [str(c).strip() for c in fila]
        break
else:
    raise ValueError("No se encontró la fila con cabeceras esperadas")


# Posiciones de columnas
i_codigo = cabeceras.index("Número de material")
i_almacen = cabeceras.index("Alm.")
i_unidad = cabeceras.index("UMB")
i_cantidad = cabeceras.index("Libre utilización")

# Recolectar el stock total por código pero solo para W199
materiales = {}

for fila in hoja.iter_rows(min_row=3, values_only=True):  # datos desde fila 3
    try:
        codigo = str(fila[i_codigo]).strip()
        almacen = str(fila[i_almacen]).strip()
        unidad = str(fila[i_unidad]).strip()
        cantidad = int(fila[i_cantidad]) if fila[i_cantidad] is not None else 0
    except (TypeError, ValueError, IndexError):
        continue

    if codigo not in materiales:
        materiales[codigo] = {"cantidad": 0, "unidad": unidad, "ubicacion": "W199"}

    if almacen == "W199":
        materiales[codigo]["cantidad"] = cantidad
        materiales[codigo]["unidad"] = unidad

# Guardar en la base de datos
for codigo, datos in materiales.items():
    Material.objects.update_or_create(
        codigo=codigo,
        defaults={
            "nombre": f"Material {codigo}",
            "cantidad": datos["cantidad"],
            "unidad": datos["unidad"],
            "ubicacion": datos["ubicacion"],
        }
    )
    print(f"Actualizado: {codigo} → {datos['cantidad']} {datos['unidad']} en W199")

